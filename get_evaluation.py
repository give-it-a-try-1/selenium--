from urllib.request import urlretrieve

from openpyxl import Workbook
from selenium.webdriver.common.by import By
# WebDriverWait 库，负责循环等待
from selenium.webdriver.support.ui import WebDriverWait
# expected_conditions 类，负责条件出发
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
import openpyxl
import json
import xlwt
import xlrd
import csv
import time
import re
import os

brower = webdriver.Chrome()
book = openpyxl.load_workbook('../景区信息.xlsx')
sheet1 = book['Sheet1']
len_row = sheet1.max_row
url = 'https://bj.meituan.com/'
login_url = 'https://passport.meituan.com/account/unitivelogin?service=www'
lost_url = 'https://www.meituan.com/error/403'
lost_list = []

max_wait_time = 10
# now_sum = 190
now_sum = 0
# 321
# 420

def add_cookies():
    brower.get(url)
    with open('../data/cookies.json', 'r', encoding='utf-8') as file:
        a = file.read()
        a = json.loads(a)
        for item in a:
            brower.add_cookie(item)
        time.sleep(1)
        # start = '广元市千佛崖石刻艺术博物馆'
        # brower.get(url + 's/' + start + '/')
        brower.get(url)
        time.sleep(1)


def login(location_original_name):
    # // *[ @ id = "qlogin_list"] / a
    # // *[ @ id = "qlogin_list"] / a
    # login_agreement = brower.find_elements_by_xpath('//*[@id="user-agreement-wrap-text-circle"]')
    # login_QQ = brower.find_elements_by_xpath('//*[@id="J-third-tencent"]')
    # login_QQ_head = brower.find_elements_by_xpath('//*[@id="qlogin_list"]')
    # # login_agreement = WebDriverWait(brower, max_wait_time).until(
    # #     EC.presence_of_all_elements_located((By.XPATH, '//*[@id="user-agreement-wrap-text-circle"]')))
    # # login_QQ = WebDriverWait(brower, max_wait_time).until(
    # #     EC.presence_of_all_elements_located((By.XPATH, '//*[@id="J-third-tencent"]')))
    # # login_QQ_head = WebDriverWait(brower, max_wait_time).until(
    # #     EC.presence_of_all_elements_located((By.XPATH, '//*[@id="qlogin_list"]')))
    # for item in login_agreement:
    #     item.click()
    #     break
    # time.sleep(1)
    # for item in login_QQ:
    #     item.click()
    #     break
    # time.sleep(1)
    # for item in login_QQ_head:
    #     item.click()
    #     break
    # time.sleep(1)
    global brower
    handles = brower.window_handles
    brower_len = len(handles)
    for i in range(0,brower_len):
        brower.switch_to.window(brower.window_handles[i])  # 切换
        brower.close()
    time.sleep(10)
    brower = webdriver.Chrome()
    lost_list.append(location_original_name)
    print(lost_list)
    add_cookies()


def except_solve():
    handles = brower.window_handles
    brower_len = len(handles)
    for i in range(1,brower_len):
        brower.switch_to.window(brower.window_handles[i])  # 切换
        brower.close()
    brower.switch_to.window(brower.window_handles[0])


def get_location():
    location_now = 0
    print(len_row)
    for num in range(2, len_row+1):
        location_original_name = sheet1.cell(num, 1).value
        location_city = sheet1.cell(num,2).value
        location_now += 1
        if location_now <= now_sum:
            continue
        # print(location_original_name)
        # if location_original_name != '青城山':
        #     continue
        print(location_original_name)
        print(location_city)

        brower.get(url + 's/' + location_original_name + '/')
        time.sleep(1)
        now_url = brower.current_url
        print(now_url)
        if (len(now_url) >= 61 and now_url[:61] == login_url) or  now_url == lost_url:
            # print(location_now)
            # add_cookies()
        # brower.get()
            login(location_original_name)

        # get_basis_information(location_original_name)
        time.sleep(3)
        try:
            get_basis_information(location_original_name,location_city)
        except Exception as e:
            print(e)
            print('失败')
            lost_list.append(location_original_name)
            print(lost_list)
            except_solve()
            time.sleep(5)
        # print(location)


def save_img(img_url, location_name):
    urlretrieve(img_url, '../data/景区图片/' + location_name + '.jpg')


def save_basis_information(map):
    # print(map)
    global now_sum
    for num in range(2, len_row):
        location_original_name = sheet1.cell(num, 1).value
        # print(location_original_name)
        if location_original_name == map['location_original_name']:
            sheet1.cell(num, 8).value = map['location_name']
            sheet1.cell(num, 9).value = map['location_tag']
            sheet1.cell(num, 10).value = map['location_score']
            sheet1.cell(num, 11).value = map['location_evaluation_num']
            sheet1.cell(num, 12).value = map['location_price']
    book.save(filename="../景区信息.xlsx")
    now_sum += 1
    print(now_sum)


def save_comment_information(path,map,flag,num,workbook,booksheet):
    # book1 = openpyxl.load_workbook(path)
    # sheet2 = book['Sheet']
    # print(sheet2)
    if flag == True:
        # sheet2.cell(1, 1).value = '评论时间'
        # sheet2.cell(1, 2).value = '评价分数'
        # sheet2.cell(1, 3).value = '评论详细'
        # sheet2.write()

        booksheet.write(0, 0, '评论时间')
        booksheet.write(0, 1, '评价分数')
        booksheet.write(0, 2, '评论详细')

    # print(map['comment_date'])
    booksheet.write(num, 0, map['comment_date'])
    booksheet.write(num, 1, map['comment_star'])
    booksheet.write(num, 2, map['comment_detail'])
    # sheet2.cell(num+1, 1).value = map['comment_date']
    # sheet2.cell(num+1, 2).value = map['comment_star']
    # sheet2.cell(num+1, 3).value = map['comment_detail']
    # print(sheet2.cell(num+1, 3).value)
    # book1.save(filename=path)

    # with open(path, 'a', encoding='utf-8-sig') as csvfile:
    #     file = csv.writer(csvfile)
    #     if flag == True:
    #         file.writerow(['评论时间','评分','评论内容'])
    #     print(map['comment_detail'])
    #     file.writerow([map['comment_date'], map['comment_star'], map['comment_detail']])


def get_basis_information(location_name,location_city):
    # locations = brower.find_elements_by_xpath('//*[@class="common-list-main"]')
    # click_locations = brower.find_elements_by_xpath('//*[@class="common-list-main"]//*[@class="default-list-item clearfix"]/a')
    # img_locations = brower.find_elements_by_xpath('//*[@class="common-list-main"]//*[@class="default-list-item clearfix"]/a/img')
    # brower.execute_script('window.scrollTo(0,document.body.scrollHeight)')  # 页面拉倒最底下
    basis_locations = WebDriverWait(brower, max_wait_time).until(
        EC.presence_of_all_elements_located((By.XPATH, '//*[@class="common-list-main"]')))
    print('one')
    img_locations = brower.find_elements_by_xpath('//*[@class="common-list-main"]//*[@class="default-list-item clearfix"]/a/img')
    print('two')
    # click_locations = brower.find_elements_by_xpath('//*[@class="common-list-main"]//*[@class="default-list-item clearfix"]/a')
    # click_locations = WebDriverWait(brower, max_wait_time).until(EC.presence_of_all_elements_located(
    #     (By.XPATH, '//*[@class="common-list-main"]//*[@class="default-list-item clearfix"]/a')))

    # img_locations = WebDriverWait(brower, max_wait_time).until(EC.presence_of_all_elements_located(
    #     (By.XPATH, '//*[@class="common-list-main"]//*[@class="default-list-item clearfix"]/a/img')))
    print('three')
    # print(locations)
    img_save_name = ''
    for item in basis_locations:
        List = item.text.split()
        # print(List)
        if List[0] == '对不起，没有符合条件的商家':
            lost_list.append(location_name)
            print(lost_list)
            break

        idea_tag = '(.*?)|.*?'
        idea_evaluation = '(.*?)分(.*?)人.*?'

        result_tag = re.findall(idea_tag, List[3], re.S)
        result_evaluation = re.findall(idea_evaluation, List[2], re.S)

        name_tag = ''
        img_save_name = List[1]
        for i in result_tag:
            if i != '|':
                name_tag += i
            else:
                break

        # print(result_tag)
        # print(name_tag)
        # print(result_evaluation)
        # location_score = (str)(result_evaluation[0] + '分')
        # print(result_evaluation[0][0])

        # print(item.text)
        map = {
            'location_original_name': location_name,
            'location_name': List[1],
            'location_score': result_evaluation[0][0] + '分',
            'location_evaluation_num': result_evaluation[0][1],
            # 'location_tag': result_tag[0],
            'location_tag': name_tag,
            'location_price': List[5]
        }
        save_basis_information(map)
        break

    for item in img_locations:
        img_url = item.get_attribute('src')
        # print(item.get_attribute('src'))
        save_img(img_url, img_save_name)
        break

    # for item in click_locations:
    #     item.click()
    #     get_comments_information(location_name,location_city)
    #     break

    time.sleep(1)


def get_comments_information(location_name,location_city):
    # brower.implicitly_wait(1) #隐式等待
    brower.switch_to.window(brower.window_handles[1])  # 切换窗口
    # comments = brower.find_elements_by_xpath('//*[@class="comment-main"]')

    page_num = 0
    comment_sum = 0
    csv_name = True
    flag = True
    workbook = xlwt.Workbook(encoding='utf-8')
    booksheet = workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
    try:
        while flag:
            flag = False
            # brower.execute_script('window.scrollTo(0,document.body.scrollHeight)')  # 页面拉倒最底下
            # read_button = WebDriverWait(brower, max_wait_time).until(
            #     EC.presence_of_all_elements_located((By.XPATH, '//*[@class="read-btn"]')))
            # next_button = brower.find_elements_by_xpath('//*[@class="pagination-item pagination-item-comment next-btn active"]')
            page_num += 1
            if page_num > 30:
                break
            comments = WebDriverWait(brower, max_wait_time).until(
                EC.presence_of_all_elements_located((By.XPATH, '//*[@class="comment-main"]')))
            print('four')
            star = WebDriverWait(brower, max_wait_time).until(
                EC.presence_of_all_elements_located((By.XPATH, '//*[@class="comment-main"]//*[@class="rate-stars-ul rate-stars-light"]')))
            print('five')
            next_button = WebDriverWait(brower, max_wait_time).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH, '//*[@class="pagination-item pagination-item-comment next-btn active"]')))
            print('six')
            read_button = brower.find_elements_by_xpath('//*[@class="read-btn"]')
            print('seven')

            for item in read_button:
                item.click()

            detail_comments = brower.find_elements_by_xpath('//*[@class="user-comment"]')
            detail_comments_date = brower.find_elements_by_xpath('//*[@class="comment-date"]')

            detail_comments_list = []
            detail_comments_date_list = []
            star_list = []

            # print(detail_comments)

            for item in detail_comments:
                # List = item.text.split()
                # print(item.text)
                comment = item.text
                detail_comments_list.append(comment)
                # print(comment)

            for item in detail_comments_date:
                date = item.text
                detail_comments_date_list.append(date)
                # print(date)

            for item in star:
                star_num = item.get_attribute('style')
                # star_list.append()
                flag1 = False
                star_num_1 = ''
                for i in star_num:
                    if i == '%':
                        break
                    if flag1 == True:
                        star_num_1 += i
                    if i == ' ':
                        flag1 = True
                # print(star_num_1)
                num = (int(star_num_1)) / 100 * 5
                # print(num)
                star_list.append(num)
                # print(star_num)

            # if csv_name == True:
            #     # wb = Workbook()
            #     wb = xlwt.Workbook(encoding="utf-8", style_compression=0)
            #     # ws = wb.active  # 获取默认sheet
            #     # wb.save('../data/景区评论/' + location_name + '.xlsx')
            #
            #     ws = wb.add_sheet('Sheet1', cell_overwrite_ok=True)
            #     wb.save('../data/景区评论/' + location_name + '.xls')


            for i in range(len(detail_comments)):
                comment_sum += 1
                map = {
                    'comment_detail' : detail_comments_list[i],
                    'comment_date' : detail_comments_date_list[i],
                    'comment_star' : star_list[i]
                }
                # save_comment_information('../data/景区评论/' + location_name + '.xlsx',map,csv_name,comment_sum)
                save_comment_information('../data/景区评论/' + location_name + '.xls', map, csv_name, comment_sum,workbook,booksheet)
                csv_name = False

            for item in next_button:
                item.click()
                flag = True

            time.sleep(1)
    except Exception as e:
        print(e)
    finally:
        workbook.save('../data/景区评论/' + location_city + '/' + location_name + '.xls')

    time.sleep(1)
    brower.close()
    brower.switch_to.window(brower.window_handles[0])  # 切换


add_cookies()
get_location()
brower.close()
print(lost_list)

